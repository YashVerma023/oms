"""Sheet upload -> MySQL loader.

Handles the three uploadable sources:

    all-users    All User.xlsx, 'Main' sheet   -> all_users     (replace)
    running      running-users.csv             -> running_users (append)
    usersetting  USERSETTING.csv, header row 7 -> usersetting   (replace)

Rules:
  * Only columns that exist in the target table are read; extra sheet columns
    are dropped.
  * Values are coerced to the column's MySQL type. Placeholders such as '#N/A',
    'NA', '-' and unparseable numbers become NULL and are counted in the report.
  * Rows without a primary key are skipped and reported.
  * Replace loads run in a single transaction: a failure leaves existing data
    untouched.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO, Callable

from sqlalchemy import bindparam, text

from core import derive, rules
from database.db import db, get_config

logger = logging.getLogger(__name__)

# Rows inserted per executemany batch.
CHUNK = 500

# Values that mean "no value" in any column: blanks and Excel error cells.
# Deliberately excludes text like 'Not Running' or 'NA', which are real values
# in columns such as `server` and `Operator Name`. Non-numeric text landing in a
# numeric column is caught by the coercers instead and reported as unparseable.
NULL_TOKENS = {
    "", "-", "--", "#n/a", "#value!", "#ref!", "#div/0!", "#name?", "#null!", "nan",
}

TRUE_TOKENS = {"true", "yes", "y", "1", "t"}
FALSE_TOKENS = {"false", "no", "n", "0", "f"}

# Usersetting files are named per server: "VS1 19 AUG 2026 USERSETTING.csv".
# The server must be the first token and separated from what follows, so a
# run-together name like "VS2819AUG2026USERSETTINGS" is rejected rather than
# silently read as server VS2819.
SERVER_IN_FILENAME = re.compile(r"^(VS\d+)(?:[\s_-]|$)", re.IGNORECASE)


def server_from_filename(filename: str) -> dict[str, Any]:
    """Extract the server code from a usersetting file name.

    Raises:
        ValueError: the name does not begin with a separated VS<number> token.
    """
    stem = filename.rsplit(".", 1)[0].strip()
    match = SERVER_IN_FILENAME.match(stem)
    if not match:
        raise ValueError(
            f"'{filename}': the file name must start with the server followed by a "
            f"space, e.g. 'VS1 19 AUG 2026 USERSETTING.csv'. "
            f"Run-together names like 'VS2819AUG2026USERSETTINGS' are not accepted."
        )
    return {"server": match.group(1).upper()}


@dataclass
class ImportSpec:
    table: str
    title: str
    kind: str                       # "xlsx" or "csv"
    mode: str                       # "replace" or "append"
    pk: tuple[str, ...]             # one or more columns forming the unique key
    header_row: int = 1             # 1-based row holding the headers
    sheet: str | None = None        # xlsx only
    accept: tuple[str, ...] = (".csv",)
    multiple: bool = False          # allow several files in one upload
    # db column -> sheet header, for the few that differ.
    renames: dict[str, str] = field(default_factory=dict)
    # Optional per-row normaliser, applied after coercion.
    post_process: Callable[[dict[str, Any]], dict[str, Any]] | None = None
    # Optional: derive extra column values from the file's name. Raising
    # ValueError rejects the upload.
    filename_columns: Callable[[str], dict[str, Any]] | None = None
    # For mode="replace_scope": only rows sharing this column's uploaded
    # values are deleted before insert.
    scope_column: str | None = None
    # Optional: run after a successful write, e.g. to derive columns that
    # come from another table.
    post_write: Callable[[], Any] | None = None
    # Optional: the upload form asks for a date and stamps it on every row.
    date_column: str | None = None
    # Rows whose column value matches one of these are dropped, e.g. the
    # trailing Total / Grand Total lines of a sheet. Matched case-insensitively.
    exclude_values: dict[str, tuple[str, ...]] = field(default_factory=dict)

    @property
    def pk_label(self) -> str:
        return " + ".join(self.pk)


IMPORT_SPECS: dict[str, ImportSpec] = {
    "all-users": ImportSpec(
        table="all_users",
        title="All Users",
        kind="xlsx",
        mode="replace_scope",
        pk=("userId",),
        header_row=1,
        sheet="Main",
        accept=(".xlsx", ".xlsm"),
        renames={
            "ml_pct": "%",
            "0SL": "SL",
            "Remarks": "Remarks/Algo8 Previous day realised MTM",
        },
        # ml_pct is recomputed and the NOT RUNNING / DLR ACC linkage enforced,
        # so the sheet's own '%' column is only a fallback.
        post_process=derive.apply_all_users,
        # The sheet has no date: the upload form asks for one, and only that
        # date's rows are replaced so earlier days are kept.
        date_column="Date",
        scope_column="Date",
    ),
    "jainam": ImportSpec(
        table="jainam",
        title="Jainam",
        kind="xlsx",
        mode="replace",
        # One row per user per date.
        pk=("Date", "UserID"),
        header_row=1,
        sheet="Jainam",
        accept=(".xlsx", ".xlsm"),
        # The sheet ends with a Total line that is not an account.
        exclude_values={"UserID": ("total", "grand total", "grandtotal")},
    ),
    "server-config": ImportSpec(
        table="server_config",
        title="Server Config",
        kind="xlsx",
        mode="replace",
        pk=("Server",),
        header_row=1,
        sheet="Servers",
        accept=(".xlsx", ".xlsm"),
        # The mapping changed, so operators must be re-derived.
        post_write=derive.sync_all_users_operator,
    ),
    "running": ImportSpec(
        table="running_users",
        title="Running Users",
        kind="csv",
        mode="append",
        pk=("userId",),
        header_row=1,
    ),
    "usersetting": ImportSpec(
        table="usersetting",
        title="Usersetting",
        kind="csv",
        # Shared accounts (FEED, dealer logins) legitimately appear in several
        # servers' files, so the account alone is not unique - the server is
        # part of the key. It comes from the file name, not the sheet.
        pk=("User ID", "server"),
        # Rows 1-6 are comment lines; row 7 holds the headers.
        header_row=7,
        # Settings arrive as one CSV per server; several can be loaded at once.
        multiple=True,
        # `server` comes from the file name, not from the sheet.
        filename_columns=server_from_filename,
        # Replace only the servers being uploaded, so loading VS1 does not
        # wipe VS2..VS28.
        mode="replace_scope",
        scope_column="server",
        # `algo` is copied from all_users once the rows are in.
        post_write=derive.sync_usersetting_algo,
    ),
}

# Columns the database fills in itself.
AUTO_COLUMNS = {"id", "created_at", "updated_at", "imported_at"}


@dataclass
class ImportReport:
    target: str
    filename: str
    loaded: int = 0
    skipped: int = 0
    issues: list[str] = field(default_factory=list)
    nulled: dict[str, int] = field(default_factory=dict)
    matched_columns: list[str] = field(default_factory=list)
    ignored_headers: list[str] = field(default_factory=list)
    files: list[dict[str, Any]] = field(default_factory=list)

    def note(self, message: str) -> None:
        # Cap the list so a badly broken file cannot flood the page or the log.
        if len(self.issues) < 50:
            self.issues.append(message)


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def _is_null(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().lower() in NULL_TOKENS
    return False


def _to_decimal(value: Any) -> Decimal | None:
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return Decimal(str(value))
    text_value = str(value).strip().replace(",", "")
    # '15%' and '0.8%' appear in percentage columns; keep the number.
    percent = text_value.endswith("%")
    text_value = text_value.rstrip("%").strip()
    try:
        number = Decimal(text_value)
    except (InvalidOperation, ValueError):
        return None
    return number / 100 if percent else number


def _to_int(value: Any) -> int | None:
    number = _to_decimal(value)
    return None if number is None else int(number)


def _to_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    token = str(value).strip().lower()
    if token in TRUE_TOKENS:
        return True
    if token in FALSE_TOKENS:
        return False
    return None


# Excel stores dates as days since 1899-12-30. Cells formatted as 'General'
# come through openpyxl as plain numbers, so a bare number in a date column is
# read as a serial. Bounded to 1900-01-01 .. 9999-12-31 so an ordinary integer
# in a genuinely numeric column is never mistaken for a date.
EXCEL_EPOCH = dt.date(1899, 12, 30)
EXCEL_SERIAL_RANGE = (1, 2958465)


def _from_excel_serial(value: Any) -> dt.datetime | None:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return None
    serial = float(value)
    if not EXCEL_SERIAL_RANGE[0] <= serial <= EXCEL_SERIAL_RANGE[1]:
        return None
    return dt.datetime.combine(EXCEL_EPOCH, dt.time.min) + dt.timedelta(days=serial)


def _to_datetime(value: Any) -> dt.datetime | None:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time.min)

    serial = _from_excel_serial(value)
    if serial is not None:
        return serial

    token = str(value).strip()
    try:
        # '2026-08-19T03:05:55.368Z' -> naive UTC (MySQL DATETIME is tz-less).
        parsed = dt.datetime.fromisoformat(token.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed


def _to_date(value: Any) -> dt.date | None:
    parsed = _to_datetime(value)
    return None if parsed is None else parsed.date()


def _to_time(value: Any) -> str | None:
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, dt.timedelta):
        return str(value)
    token = str(value).strip()
    return token if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", token) else None


def _to_text(value: Any, limit: int) -> str:
    token = str(value).strip()
    return token[:limit] if limit and len(token) > limit else token


def _fit_decimal(number: Decimal | None, column: dict[str, Any]) -> Decimal | None:
    """Round to the column's scale and reject values wider than its precision.

    Without this a single oversized cell raises 'Out of range value' and aborts
    the whole upload; instead that one value becomes NULL and is reported.
    """
    if number is None:
        return None

    precision, scale = column["precision"], column["scale"]
    if not precision:
        return number

    rounded = number.quantize(Decimal(1).scaleb(-scale))
    if len(rounded.as_tuple().digits) - scale > precision - scale:
        return None
    return rounded


def _coerce(value: Any, column: dict[str, Any]) -> Any:
    """Convert one sheet value to something MySQL will accept for this column."""
    if _is_null(value):
        return None

    data_type = column["data_type"]
    if column["is_bool"]:
        return _to_bool(value)
    if data_type == "decimal":
        return _fit_decimal(_to_decimal(value), column)
    if data_type in ("float", "double"):
        return _to_decimal(value)
    if data_type in ("int", "bigint", "smallint", "mediumint", "tinyint"):
        return _to_int(value)
    if data_type == "datetime":
        return _to_datetime(value)
    if data_type == "date":
        return _to_date(value)
    if data_type == "time":
        return _to_time(value)
    return _to_text(value, column["length"])


# ---------------------------------------------------------------------------
# Target metadata
# ---------------------------------------------------------------------------

def _target_columns(table: str) -> list[dict[str, Any]]:
    """Writable columns of the target table with everything needed to coerce."""
    rows = db.session.execute(
        text(
            "SELECT COLUMN_NAME, DATA_TYPE, COLUMN_TYPE, CHARACTER_MAXIMUM_LENGTH, "
            "NUMERIC_PRECISION, NUMERIC_SCALE "
            "FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA = :schema AND TABLE_NAME = :table "
            "ORDER BY ORDINAL_POSITION"
        ),
        {"schema": get_config()["database"], "table": table},
    ).all()

    return [
        {
            "name": name,
            "data_type": data_type,
            # MySQL BOOLEAN is an alias for tinyint(1).
            "is_bool": column_type.lower().startswith("tinyint(1)"),
            "length": int(length) if length else 0,
            "precision": int(precision) if precision else 0,
            "scale": int(scale) if scale else 0,
        }
        for name, data_type, column_type, length, precision, scale in rows
        if name not in AUTO_COLUMNS
    ]


# ---------------------------------------------------------------------------
# Sheet reading
# ---------------------------------------------------------------------------

def _read_csv(stream: BinaryIO, header_row: int) -> tuple[list[str], list[list[str]]]:
    raw = stream.read()
    try:
        decoded = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        logger.warning("File is not UTF-8; falling back to latin-1")
        decoded = raw.decode("latin-1")

    rows = list(csv.reader(io.StringIO(decoded)))
    if len(rows) < header_row:
        raise ValueError(f"File has {len(rows)} rows but headers were expected on row {header_row}")

    headers = [h.strip() for h in rows[header_row - 1]]
    return headers, rows[header_row:]


def _read_xlsx(stream: BinaryIO, sheet: str, header_row: int) -> tuple[list[str], list[list[Any]]]:
    import openpyxl  # imported lazily: only xlsx uploads need it

    # data_only=True reads the cached result of formulas rather than '=VLOOKUP(...)'.
    book = openpyxl.load_workbook(io.BytesIO(stream.read()), data_only=True, read_only=True)
    if sheet not in book.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {', '.join(book.sheetnames)}")

    rows = list(book[sheet].iter_rows(values_only=True))
    book.close()

    if len(rows) < header_row:
        raise ValueError(f"Sheet has {len(rows)} rows but headers were expected on row {header_row}")

    headers = [str(h).strip() if h is not None else "" for h in rows[header_row - 1]]
    return headers, [list(r) for r in rows[header_row:]]


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def _parse_one(
    spec: ImportSpec,
    columns: list[dict[str, Any]],
    stream: BinaryIO,
    filename: str,
    report: ImportReport,
    label: str,
    extra: dict[str, Any],
) -> tuple[list[tuple[dict[str, Any], int]], list[tuple[str, dict[str, Any]]]]:
    """Parse one file into (matched columns, [(primary key, record), ...])."""
    # Rejects the file before it is parsed if the name is not usable.
    from_name = spec.filename_columns(filename) if spec.filename_columns else {}
    # Values supplied by the upload form (e.g. the chosen date) win over the
    # sheet, and are stamped on every row.
    supplied = {**from_name, **extra}

    if spec.kind == "xlsx":
        headers, data_rows = _read_xlsx(stream, spec.sheet, spec.header_row)
    else:
        headers, data_rows = _read_csv(stream, spec.header_row)

    header_index = {h: i for i, h in enumerate(headers) if h}

    # Match each table column to a sheet column, honouring the rename map.
    matched: list[tuple[dict[str, Any], int]] = []
    for column in columns:
        if column["name"] in supplied:
            continue  # supplied by the form or file name, not the sheet
        source_header = spec.renames.get(column["name"], column["name"])
        if source_header in header_index:
            matched.append((column, header_index[source_header]))
        else:
            report.note(f"{label}Column '{column['name']}' not found in the sheet - loaded as NULL")

    if not matched:
        raise ValueError(
            f"{filename}: no table columns matched the sheet headers. "
            f"Headers found: {', '.join(h for h in headers[:12] if h)}..."
        )

    positions = {c["name"]: i for c, i in matched}
    # Part of the key can come from the file name or the form rather than the
    # sheet - usersetting is keyed on User ID + server, and server is in the
    # file name. Only the rest has to be present as a column.
    sheet_pk = [name for name in spec.pk if name not in supplied]
    missing_pk = [name for name in sheet_pk if name not in positions]
    if missing_pk:
        raise ValueError(
            f"{filename}: key column(s) {', '.join(missing_pk)} missing from the sheet"
        )
    pk_positions = [positions[name] for name in sheet_pk]

    used = {i for _, i in matched}
    for header in (h for i, h in enumerate(headers) if h and i not in used):
        if header not in report.ignored_headers:
            report.ignored_headers.append(header)

    parsed: list[tuple[str, dict[str, Any]]] = []

    for offset, raw_row in enumerate(data_rows):
        sheet_row = spec.header_row + offset + 1

        if all(_is_null(v) for v in raw_row):
            continue  # blank spacer row

        if _excluded(spec, matched, raw_row):
            report.skipped += 1
            report.note(f"{label}Row {sheet_row}: summary row - skipped")
            continue

        key_parts = [
            raw_row[position] if position < len(raw_row) else None
            for position in pk_positions
        ]
        if any(_is_null(part) for part in key_parts):
            report.skipped += 1
            report.note(f"{label}Row {sheet_row}: missing {spec.pk_label} - skipped")
            continue

        # Keyed by real column name so post_process rules stay readable.
        values: dict[str, Any] = dict(supplied)
        for column, index in matched:
            raw = raw_row[index] if index < len(raw_row) else None
            value = _coerce(raw, column)
            if value is None and not _is_null(raw):
                report.nulled[column["name"]] = report.nulled.get(column["name"], 0) + 1
            values[column["name"]] = value

        if spec.post_process:
            spec.post_process(values)
            # A rule may have produced a value that must still fit its column.
            for column, _ in matched:
                if column["data_type"] == "decimal":
                    values[column["name"]] = _fit_decimal(values[column["name"]], column)

        # Built from the assembled row, so key parts that came from the file
        # name count too.
        key = "\x1f".join(str(values.get(name, "")).strip() for name in spec.pk)
        parsed.append((key, {_param(name): value for name, value in values.items()}))

    return matched, parsed


def import_sheet(
    target: str,
    uploads: list[tuple[BinaryIO, str]],
    extra: dict[str, Any] | None = None,
) -> ImportReport:
    """Parse one or more uploaded sheets and load them into the target table.

    Multiple files are parsed first and written as a single transaction, so a
    replace load swaps in the combined contents of every file at once rather
    than each file wiping the previous one. Within a replace load the last
    occurrence of a primary key wins, across files as well as within one.

    Raises:
        KeyError: unknown target.
        ValueError: unreadable file, wrong sheet, or no usable columns.
    """
    spec = IMPORT_SPECS[target]
    filenames = [name for _, name in uploads]
    report = ImportReport(target=target, filename=", ".join(filenames))

    columns = _target_columns(spec.table)
    if not columns:
        raise ValueError(f"Table '{spec.table}' does not exist - restart the app to provision it")

    matched: list[tuple[dict[str, Any], int]] = []
    collected: list[tuple[str, str, dict[str, Any]]] = []   # (file, key, record)

    for stream, filename in uploads:
        # Only prefix notes with the filename when there is more than one.
        label = f"{filename}: " if len(uploads) > 1 else ""
        matched, parsed = _parse_one(
            spec, columns, stream, filename, report, label, extra or {}
        )
        collected.extend((filename, key, record) for key, record in parsed)
        report.files.append({"name": filename, "rows": len(parsed)})

    if not collected:
        raise ValueError("No loadable rows found")

    # No duplicate keys are ever loaded, for any target: the last occurrence of
    # a primary key wins, whether the repeat is inside one file or across files.
    winners: dict[str, tuple[str, dict[str, Any]]] = {}
    for filename, key, record in collected:
        if key in winners:
            report.skipped += 1
            previous = winners[key][0]
            report.note(
                f"Duplicate {spec.pk_label} '{key.replace(chr(31), ' + ')}' - kept the row "
                f"from '{filename}', dropped the one from '{previous}'"
                if previous != filename
                else f"Duplicate {spec.pk_label} '{key.replace(chr(31), ' + ')}' "
                     f"in '{filename}' - kept the last row"
            )
        winners[key] = (filename, record)

    records = [record for _, record in winners.values()]

    report.matched_columns = [c["name"] for c, _ in matched]
    _write(spec, matched, records, report)

    if spec.post_write:
        derived = spec.post_write()
        if derived:
            report.note(f"Derived columns updated on {derived} row(s) after load.")

    logger.info(
        "Import '%s' from %s file(s) [%s]: %s loaded, %s skipped, %s columns matched",
        target, len(uploads), report.filename, report.loaded, report.skipped, len(matched),
    )
    return report


def _excluded(
    spec: ImportSpec,
    matched: list[tuple[dict[str, Any], int]],
    raw_row: list[Any],
) -> bool:
    """Whether this row is a summary line the spec asks to drop."""
    if not spec.exclude_values:
        return False

    positions = {c["name"]: i for c, i in matched}
    for column, unwanted in spec.exclude_values.items():
        index = positions.get(column)
        if index is None or index >= len(raw_row):
            continue
        value = str(raw_row[index] or "").strip().lower()
        if value and value in {u.lower() for u in unwanted}:
            return True
    return False


def _param(column_name: str) -> str:
    """Bind-parameter name for a column (names may contain spaces or symbols)."""
    return "p_" + re.sub(r"\W", "_", column_name)


def _write(
    spec: ImportSpec,
    matched: list[tuple[dict[str, Any], int]],
    records: list[dict[str, Any]],
    report: ImportReport,
) -> None:
    """Replace or append the parsed rows in one transaction."""
    # Records can carry columns from three places: matched sheet headers, the
    # file name, and post_process (e.g. the derived `Date`). Build the
    # param -> column map from the table itself so every one is covered.
    names = sorted(records[0].keys())
    reverse = {_param(c["name"]): c["name"] for c in _target_columns(spec.table)}

    unknown = [p for p in names if p not in reverse]
    if unknown:
        raise ValueError(
            f"Internal error: parameters {unknown} do not map to columns of "
            f"'{spec.table}'"
        )
    column_list = ", ".join(f"`{reverse[p]}`" for p in names)
    value_list = ", ".join(f":{p}" for p in names)
    insert_sql = text(f"INSERT INTO `{spec.table}` ({column_list}) VALUES ({value_list})")

    try:
        if spec.mode == "replace":
            # DELETE, not TRUNCATE: TRUNCATE commits implicitly and would
            # destroy the old data even if the insert then failed.
            deleted = db.session.execute(text(f"DELETE FROM `{spec.table}`")).rowcount
            logger.info("Replace mode: removed %s existing rows from '%s'", deleted, spec.table)

        elif spec.mode == "replace_scope":
            # Only the uploaded scopes are cleared, so loading one server's
            # file leaves every other server untouched.
            scope_param = _param(spec.scope_column)
            scopes = sorted({r[scope_param] for r in records if r.get(scope_param) is not None})
            if scopes:
                deleted = db.session.execute(
                    text(
                        f"DELETE FROM `{spec.table}` "
                        f"WHERE `{spec.scope_column}` IN :scopes"
                    ).bindparams(bindparam("scopes", expanding=True)),
                    {"scopes": scopes},
                ).rowcount
                # Scope values are not always strings - `Date` is a date.
                shown = ", ".join(str(s) for s in scopes)
                logger.info(
                    "Scoped replace: removed %s row(s) from '%s' for %s %s",
                    deleted, spec.table, spec.scope_column, shown,
                )
                report.note(
                    f"Replaced existing rows for {spec.scope_column} {shown} "
                    f"({deleted} removed); other {spec.scope_column} values "
                    f"were left untouched."
                )

        for start in range(0, len(records), CHUNK):
            db.session.execute(insert_sql, records[start:start + CHUNK])

        db.session.commit()
        report.loaded = len(records)
    except Exception:
        db.session.rollback()
        logger.exception("Import into '%s' failed; no changes were written", spec.table)
        raise
